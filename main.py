from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from openpyxl import load_workbook

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Autorise toutes les origines (changer pour prod)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = "PORTALIA_MC2_CONSULTANTS_2024_V03-24.xlsm"  # Remplace par le chemin correct
SHEET_NAME = "1. Calcul Avec prov"  # Remplace par le bon nom de feuille

@app.post("/calculate")
def calculate_salary(data: dict):
    try:
        # Charger le fichier Excel
        wb = load_workbook(EXCEL_FILE, keep_vba=True)
        ws = wb[SHEET_NAME]

        # Insérer les valeurs dans la colonne J selon l'image fournie
        ws["J4"] = data.get("tjm", None)  # Taux Journalier Moyen
        ws["J5"] = data.get("joursTravailles", 18)  # Nombre de jours travaillés
        ws["J8"] = data.get("cdi_2", 0.02)  # CDI - 2%
        ws["J9"] = data.get("cdi_nego", "A négocier")  # CDI - négociation
        ws["J10"] = data.get("cdi_autre", 0.00)  # Autre paramètre CDI
        ws["J12"] = data.get("frais_fonctionnement", "A négocier")  # Frais de fonctionnement
        ws["J21"] = data.get("ticket_resto", 198)  # Ticket restaurant
        ws["J17"] = data.get("mutuelle", "Oui")  # Mutuelle
        ws["J25"] = data.get("code_commune", "Nombre donné")  # Code commune

        # Sauvegarder et recharger pour activer les calculs Excel
        wb.save(EXCEL_FILE)
        wb.close()

        # Recharger le fichier pour lire les résultats
        df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME, engine="openpyxl")

        # Lire les résultats depuis la bonne colonne (adapte si besoin)
        results = {
            "tjm": df.at[3, "C"],  # Ligne 4, colonne C (à adapter selon l'Excel)
            "brut": df.at[5, "B"],  # Ligne 5, colonne C
            "net": df.at[9, "B"],  # Ligne 6, colonne C
        }

        return results

    except Exception as e:
        return {"error": str(e)}
